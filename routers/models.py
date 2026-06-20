import os
import uuid
import json
import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import JSONResponse, RedirectResponse, Response

import db.factory as fdb
import db.users as udb
import auth
from config import IMAGES_DIR, CURRENCIES

router = APIRouter(prefix="/models")
from tmpl import templates


def _options_for_category(category_id, mfr_filter=None):
    """Return options matching the given category (and optionally manufacturer)."""
    all_opts = fdb.get_options(manufacturer_filter=mfr_filter)
    if not category_id:
        return all_opts
    cid = str(category_id)
    result = []
    for o in all_opts:
        ids = [x.strip() for x in (o.get("category_ids") or "").split(",") if x.strip()]
        if not ids or cid in ids:
            result.append(o)
    return result


def _cats_for_user(user):
    """Return categories filtered by user's allowed_categories (manufacturer isolation)."""
    cats = fdb.get_cats()
    if not auth.is_mfr(user):
        return cats
    allowed_names = set(x.strip() for x in (user.get("allowed_categories") or "").split(",") if x.strip())
    if not allowed_names:
        return cats
    return [c for c in cats if c["name"] in allowed_names]


def _calc_cost(purchase_price, shipping_cost, customs_pct, extra_tax_pct,
               port_cost, document_cost, installation_cost, other_cost):
    return (
        purchase_price + shipping_cost
        + purchase_price * customs_pct / 100
        + purchase_price * extra_tax_pct / 100
        + port_cost + document_cost + installation_cost + other_cost
    )


@router.get("")
async def models_list(request: Request, category_id: int = 0):
    user = auth.require_user(request)
    cats = fdb.get_cats()
    models = fdb.get_models(category_id if category_id else None)
    if auth.is_mfr(user):
        allowed_ids = set(int(x) for x in (user.get("allowed_models") or "").split(",") if x.strip().isdigit())
        group_ids = udb.get_manufacturer_group_ids(user)
        # allowed_categories stored as names (e.g. "CNC İşleme,PVC Kenar Bantlama")
        allowed_cat_names = set(x.strip() for x in (user.get("allowed_categories") or "").split(",") if x.strip())
        if allowed_cat_names:
            # convert allowed names → id set using cats list
            allowed_cat_ids = {c["id"] for c in cats if c["name"] in allowed_cat_names}
            cats = [c for c in cats if c["id"] in allowed_cat_ids]
        else:
            allowed_cat_ids = set()
        models = [
            m for m in models
            if (m["id"] in allowed_ids or m.get("manufacturer_id") in group_ids)
            and (not allowed_cat_ids or m.get("category_id") in allowed_cat_ids)
        ]
    cat_map = {
        c["id"]: (c.get(f"name_{user.get('lang','tr')}") or c["name"])
        if user.get("lang") and user.get("lang") != "tr" else c["name"]
        for c in cats
    }
    for m in models:
        m["category_name"] = cat_map.get(m.get("category_id"), "-")
    return templates.TemplateResponse(request, "models.html", {
        "user": user,
        "models": models,
        "categories": cats,
        "selected_cat": category_id,
        "currencies": CURRENCIES,
        "active_page": "models",
    })


@router.get("/new")
async def model_new(request: Request):
    user = auth.require_user(request)
    cats = _cats_for_user(user)
    _mfr = udb.get_manufacturer_group_ids(user) if auth.is_mfr(user) else None
    options = fdb.get_options(manufacturer_filter=_mfr)
    manufacturers = udb.all_manufacturers() if user["role"] == "admin" else []
    return templates.TemplateResponse(request, "model_form.html", {
        "user": user,
        "is_mfr": auth.is_mfr(user),
        "model": {},
        "categories": cats,
        "options": options,
        "compatible_options_selected": [],
        "currencies": CURRENCIES,
        "manufacturers": manufacturers,
        "active_page": "models",
        "line_images": {},
    })


@router.get("/{model_id}/copy")
async def model_copy(request: Request, model_id: int):
    user = auth.require_user(request)
    m = fdb.get_model(model_id)
    if not m:
        return RedirectResponse("/models", 303)
    cats = _cats_for_user(user)
    _mfr = udb.get_manufacturer_group_ids(user) if auth.is_mfr(user) else None
    options = _options_for_category(m.get("category_id"), mfr_filter=_mfr)
    manufacturers = udb.all_manufacturers() if user["role"] == "admin" else []
    compatible = []
    if m.get("compatible_options"):
        try:
            compatible = json.loads(m["compatible_options"])
        except Exception:
            pass
    m_copy = dict(m)
    m_copy["id"] = 0
    m_copy["name"] = m["name"] + " (Kopya)"
    return templates.TemplateResponse(request, "model_form.html", {
        "user": user,
        "is_mfr": auth.is_mfr(user),
        "model": m_copy,
        "categories": cats,
        "options": options,
        "compatible_options_selected": compatible,
        "currencies": CURRENCIES,
        "manufacturers": manufacturers,
        "active_page": "models",
        "line_images": {},
    })


@router.get("/{model_id}/edit")
async def model_edit(request: Request, model_id: int):
    user = auth.require_user(request)
    m = fdb.get_model(model_id)
    if not m:
        return RedirectResponse("/models", 303)
    if auth.is_mfr(user):
        allowed_ids = set(int(x) for x in (user.get("allowed_models") or "").split(",") if x.strip().isdigit())
        group_ids = udb.get_manufacturer_group_ids(user)
        if model_id not in allowed_ids and m.get("manufacturer_id") not in group_ids:
            return RedirectResponse("/models", 303)
    cats = _cats_for_user(user)
    _mfr = udb.get_manufacturer_group_ids(user) if auth.is_mfr(user) else None
    options = _options_for_category(m.get("category_id"), mfr_filter=_mfr)
    manufacturers = udb.all_manufacturers() if user["role"] == "admin" else []
    compatible = []
    if m.get("compatible_options"):
        try:
            compatible = json.loads(m["compatible_options"])
        except Exception:
            pass
    line_images = fdb.get_model_line_images(model_id)
    return templates.TemplateResponse(request, "model_form.html", {
        "user": user,
        "is_mfr": auth.is_mfr(user),
        "model": m,
        "categories": cats,
        "options": options,
        "compatible_options_selected": compatible,
        "currencies": CURRENCIES,
        "manufacturers": manufacturers,
        "active_page": "models",
        "line_images": {img["line_count"]: img for img in line_images},
    })


@router.post("/save")
async def save_model(request: Request,
                     id: int = Form(0),
                     name: str = Form(...),
                     category_id: int = Form(0),
                     description: str = Form(""),
                     base_price: float = Form(0.0),
                     currency: str = Form("USD"),
                     specs: str = Form(""),
                     purchase_price: float = Form(0.0),
                     purchase_currency: str = Form("USD"),
                     shipping_cost: float = Form(0.0),
                     customs_pct: float = Form(0.0),
                     extra_tax_pct: float = Form(0.0),
                     port_cost: float = Form(0.0),
                     document_cost: float = Form(0.0),
                     installation_cost: float = Form(0.0),
                     other_cost: float = Form(0.0),
                     image: Optional[UploadFile] = File(None),
                     name_en: str = Form(""),
                     description_en: str = Form(""),
                     name_zh: str = Form(""),
                     description_zh: str = Form(""),
                     specs_en: str = Form(""),
                     specs_zh: str = Form(""),
                     is_line: int = Form(0),
                     line_configs: str = Form("2,3,4"),
                     manufacturer_id: int = Form(0),
                     catalog_sort: int = Form(999)):
    user = auth.require_user(request)

    # Manufacturers own their models; can only change purchase_price if currently 0
    if auth.is_mfr(user):
        base_price = 0.0
        shipping_cost = customs_pct = extra_tax_pct = 0.0
        port_cost = document_cost = installation_cost = other_cost = 0.0
        manufacturer_id = udb.effective_mfr_id(user)
        if id:  # editing existing — only allow price change if price not yet set
            existing = fdb.get_model(id)
            existing_price = float(existing.get("purchase_price") or 0) if existing else 0.0
            if existing_price > 0:
                # price already set — must use price request to change
                purchase_price = existing_price
                purchase_currency = existing.get("purchase_currency", "USD") if existing else "USD"
            # else: existing_price == 0 → allow manufacturer to set purchase_price from form

    total_cost = _calc_cost(
        purchase_price, shipping_cost, customs_pct, extra_tax_pct,
        port_cost, document_cost, installation_cost, other_cost,
    )

    form_data = await request.form()
    compatible_ids = form_data.getlist("compatible_options")
    compatible_json = json.dumps([int(x) for x in compatible_ids if x])

    image_path = ""
    if image and image.filename:
        os.makedirs(IMAGES_DIR, exist_ok=True)
        content = await image.read()
        fname = f"{uuid.uuid4().hex}.jpg"
        fpath = os.path.join(IMAGES_DIR, fname)
        try:
            from PIL import Image as PILImage
            import io as _io
            img = PILImage.open(_io.BytesIO(content))
            if img.mode in ("RGBA", "P", "LA"):
                img = img.convert("RGB")
            img.thumbnail((800, 800))
            img.save(fpath, "JPEG", quality=85)
        except Exception:
            # PIL failed – save raw bytes with original extension
            orig_ext = os.path.splitext(image.filename)[1].lower() or ".jpg"
            fname = f"{uuid.uuid4().hex}{orig_ext}"
            fpath = os.path.join(IMAGES_DIR, fname)
            with open(fpath, "wb") as f:
                f.write(content)
        image_path = f"img/uploads/{fname}"

    kw = dict(
        name=name,
        category_id=category_id or None,
        description=description,
        base_price=base_price,
        currency=currency,
        specs=specs,
        purchase_price=purchase_price,
        purchase_currency=purchase_currency,
        shipping_cost=shipping_cost,
        customs_pct=customs_pct,
        extra_tax_pct=extra_tax_pct,
        port_cost=port_cost,
        document_cost=document_cost,
        installation_cost=installation_cost,
        other_cost=other_cost,
        total_cost=total_cost,
        compatible_options=compatible_json,
        name_en=name_en,
        description_en=description_en,
        name_zh=name_zh,
        description_zh=description_zh,
        specs_en=specs_en,
        specs_zh=specs_zh,
        is_line=is_line,
        line_configs=line_configs.strip(),
        manufacturer_id=manufacturer_id or None,
        catalog_sort=catalog_sort if catalog_sort and catalog_sort != 999 else 999,
    )
    if image_path:
        kw["image_path"] = image_path

    if id:
        fdb.upd_model(id, **kw)
        model_id = id
    else:
        model_id = fdb.add_model(**kw)

    # Save line images
    form_data = await request.form()
    for key in form_data.keys():
        if key.startswith("line_img_count_"):
            lc = int(key.replace("line_img_count_", ""))
            prio = int(form_data.get(f"line_img_prio_{lc}", 0) or 0)
            img_file = form_data.get(f"line_img_{lc}")
            if img_file and hasattr(img_file, "filename") and img_file.filename:
                ext = os.path.splitext(img_file.filename)[1].lower() or ".jpg"
                fname = f"line_{model_id}_{lc}_{uuid.uuid4().hex[:8]}{ext}"
                fpath = os.path.join(IMAGES_DIR, fname)
                content = await img_file.read()
                try:
                    from PIL import Image as PILImage
                    import io as _io
                    img_obj = PILImage.open(_io.BytesIO(content))
                    img_obj.thumbnail((800, 800))
                    img_obj.save(fpath, "JPEG", quality=85)
                except Exception:
                    with open(fpath, "wb") as f:
                        f.write(content)
                fdb.save_model_line_image(model_id, lc, f"img/uploads/{fname}", prio)
            else:
                # Update priority only if image already exists
                existing = fdb.get_model_line_images(model_id)
                for ei in existing:
                    if ei["line_count"] == lc:
                        fdb.save_model_line_image(model_id, lc, ei["image_path"], prio)

    return RedirectResponse(f"/models/{model_id}/edit", 303)


@router.post("/spec-image")
async def upload_spec_image(request: Request, image: UploadFile = File(...)):
    auth.require_user(request)
    if not image or not image.filename:
        return JSONResponse({"error": "No file"}, status_code=400)
    os.makedirs(IMAGES_DIR, exist_ok=True)
    content = await image.read()
    fname = f"spec_{uuid.uuid4().hex[:10]}.jpg"
    fpath = os.path.join(IMAGES_DIR, fname)
    try:
        from PIL import Image as PILImage
        import io as _io
        img = PILImage.open(_io.BytesIO(content))
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        img.thumbnail((800, 800))
        img.save(fpath, "JPEG", quality=85)
    except Exception:
        orig_ext = os.path.splitext(image.filename)[1].lower() or ".jpg"
        fname = f"spec_{uuid.uuid4().hex[:10]}{orig_ext}"
        fpath = os.path.join(IMAGES_DIR, fname)
        with open(fpath, "wb") as f:
            f.write(content)
    return JSONResponse({"path": f"img/uploads/{fname}"})


@router.post("/line-image/delete")
async def delete_line_image(request: Request, id: int = Form(...), model_id: int = Form(...)):
    auth.require_user(request)
    fdb.del_model_line_image(id)
    return RedirectResponse(f"/models/{model_id}/edit", 303)


@router.get("/general-catalog-pdf")
async def general_catalog_pdf(request: Request, lang: str = "", category_id: int = 0):
    import asyncio
    import unicodedata
    from catalog_ai import generate_general_catalog_content
    from tmpl import templates

    user = auth.require_user(request)
    if not lang:
        lang = user.get("lang") or "tr"

    all_models = fdb.get_models(category_id=category_id or None)
    cats = fdb.get_cats()
    cat_map = {c["id"]: c for c in cats}
    company = fdb.get_company()

    # Attach category name + parsed specs to each model
    enriched = []
    for m in all_models:
        cat = cat_map.get(m.get("category_id"), {})
        m = dict(m)
        m["_cat_name"] = cat.get(f"name_{lang}") or cat.get("name") or ""
        specs_raw = m.get("specs") or "[]"
        try:
            m["_specs"] = json.loads(specs_raw) if isinstance(specs_raw, str) else specs_raw
        except Exception:
            m["_specs"] = []
        enriched.append(m)

    co_name = (company.get("company_name") or "") if company else ""

    # AI taglines + spec texts for all models — run in thread
    ai = await asyncio.to_thread(generate_general_catalog_content, enriched, lang, co_name)
    ai_intro        = ai.get("catalog_intro", "")
    taglines        = ai.get("model_taglines", {})
    model_spec_texts = ai.get("model_spec_texts", {})

    # Attach AI-generated spec texts to each model
    for m in enriched:
        mname = m.get(f"name_{lang}") or m.get("name", "")
        ai_st = model_spec_texts.get(mname) or model_spec_texts.get(m.get("name", "")) or {}
        m["_ai_spec_texts"] = ai_st

    _L = {
        "tr": {"product_catalog": "Ürün Kataloğu", "products": "Ürün",
               "contact_us": "İletişim", "phone": "Telefon",
               "email": "E-posta", "web": "Web", "address": "Adres",
               "specs": "Teknik Özellikler"},
        "en": {"product_catalog": "Product Catalog", "products": "Products",
               "contact_us": "Contact Us", "phone": "Phone",
               "email": "Email", "web": "Website", "address": "Address",
               "specs": "Technical Specifications"},
        "zh": {"product_catalog": "产品目录", "products": "产品",
               "contact_us": "联系我们", "phone": "电话",
               "email": "邮箱", "web": "网站", "address": "地址",
               "specs": "技术规格"},
    }
    lbl = _L.get(lang, _L["tr"])

    base_url = str(request.base_url).rstrip("/")
    date_str = datetime.date.today().strftime("%Y-%m-%d")

    html_str = templates.env.get_template("general_catalog_pdf.html").render({
        "models": enriched,
        "company": company,
        "ai_intro": ai_intro,
        "taglines": taglines,
        "lang": lang,
        "base_url": base_url,
        "date_str": date_str,
        "label_product_catalog": lbl["product_catalog"],
        "label_products":        lbl["products"],
        "label_contact_us":      lbl["contact_us"],
        "label_phone":           lbl["phone"],
        "label_email":           lbl["email"],
        "label_web":             lbl["web"],
        "label_address":         lbl["address"],
        "label_specs":           lbl["specs"],
    })

    def _render():
        from weasyprint import HTML
        return HTML(string=html_str, base_url=base_url).write_pdf()

    try:
        pdf = await asyncio.to_thread(_render)
    except Exception as e:
        return Response(f"PDF oluşturulamadı: {e}", status_code=500, media_type="text/plain")

    import unicodedata
    safe_co = unicodedata.normalize("NFKD", co_name)
    safe_co = "".join(c for c in safe_co if ord(c) < 128).replace(" ", "_")[:30] or "katalog"
    filename = f"genel_katalog_{safe_co}_{lang}.pdf"
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/{model_id}/catalog-ai-test")
async def catalog_ai_test(request: Request, model_id: int, lang: str = "tr"):
    """Debug endpoint — returns raw AI output as JSON."""
    import asyncio
    from catalog_ai import generate_catalog_content
    from config import ANTHROPIC_API_KEY
    user = auth.require_user(request)
    if user.get("role") != "admin":
        return JSONResponse({"error": "admin only"}, 403)
    m = fdb.get_model(model_id)
    if not m:
        return JSONResponse({"error": "model not found"}, 404)
    specs_raw = m.get("specs") or "[]"
    try:
        specs = json.loads(specs_raw) if isinstance(specs_raw, str) else specs_raw
    except Exception:
        specs = []
    compatible_ids = []
    try:
        compatible_ids = json.loads(m.get("compatible_options") or "[]")
    except Exception:
        pass
    options = [o for o in fdb.get_options() if o["id"] in compatible_ids]
    ai = await asyncio.to_thread(generate_catalog_content, m, lang, specs, options)
    return JSONResponse({
        "api_key_set": bool(ANTHROPIC_API_KEY),
        "api_key_prefix": ANTHROPIC_API_KEY[:12] + "..." if ANTHROPIC_API_KEY else "",
        "ai_result": ai,
    })


@router.get("/{model_id}/catalog-pdf")
async def catalog_pdf(request: Request, model_id: int, lang: str = ""):
    import asyncio
    import unicodedata
    from catalog_ai import generate_catalog_content
    from tmpl import templates

    user = auth.require_user(request)
    m = fdb.get_model(model_id)
    if not m:
        return RedirectResponse("/models", 303)

    if not lang:
        lang = user.get("lang") or "tr"

    # Parse specs
    specs_key = "specs" if lang == "tr" else f"specs_{lang}"
    raw = m.get(specs_key) or m.get("specs") or "[]"
    try:
        specs = json.loads(raw) if isinstance(raw, str) else raw
    except Exception:
        specs = []

    # Compatible options
    compatible_ids = []
    try:
        compatible_ids = json.loads(m.get("compatible_options") or "[]")
    except Exception:
        pass
    all_opts = fdb.get_options()
    options = [o for o in all_opts if o["id"] in compatible_ids]

    # Category / company / name
    cats = fdb.get_cats()
    cat = {c["id"]: c for c in cats}.get(m.get("category_id"), {})
    category_name = cat.get(f"name_{lang}") or cat.get("name") or ""
    company = fdb.get_company()
    model_name = m.get(f"name_{lang}") or m.get("name") or ""

    # AI — run in thread so it doesn't block the event loop
    ai = await asyncio.to_thread(generate_catalog_content, m, lang, specs, options)

    _L = {
        "tr": {
            "overview": "Ürün Genel Bakış",
            "highlights": "Temel Özellikler",
            "specs": "Teknik Özellikler",
            "spec_feature": "Özellik",
            "spec_detail": "Detay",
            "options": "Opsiyonlar & Aksesuarlar",
            "phone": "Telefon",
            "email": "E-posta",
            "web": "Web",
            "product_catalog": "Ürün Kataloğu",
            "contact_us": "İletişim",
            "address": "Adres",
            "option_benefits": "Seçili Opsiyonların Sağladığı Faydalar",
        },
        "en": {
            "overview": "Product Overview",
            "highlights": "Key Features",
            "specs": "Technical Specifications",
            "spec_feature": "Feature",
            "spec_detail": "Detail",
            "options": "Options & Accessories",
            "phone": "Phone",
            "email": "Email",
            "web": "Website",
            "product_catalog": "Product Catalog",
            "contact_us": "Contact Us",
            "address": "Address",
            "option_benefits": "Benefits of Selected Options",
        },
        "zh": {
            "overview": "产品概览",
            "highlights": "核心特点",
            "specs": "技术规格",
            "spec_feature": "特性",
            "spec_detail": "详情",
            "options": "选项与配件",
            "phone": "电话",
            "email": "邮箱",
            "web": "网站",
            "product_catalog": "产品目录",
            "contact_us": "联系我们",
            "address": "地址",
            "option_benefits": "所选选项的优势",
        },
    }
    lbl = _L.get(lang, _L["tr"])

    base_url = str(request.base_url).rstrip("/")
    date_str = datetime.date.today().strftime("%Y-%m-%d")

    html_str = templates.env.get_template("catalog_pdf.html").render({
        "model": m,
        "model_name": model_name,
        "specs": specs,
        "options": options,
        "category_name": category_name,
        "company": company,
        "ai": ai,
        "lang": lang,
        "base_url": base_url,
        "date_str": date_str,
        "label_overview":        lbl["overview"],
        "label_highlights":      lbl["highlights"],
        "label_specs":           lbl["specs"],
        "label_spec_feature":    lbl["spec_feature"],
        "label_spec_detail":     lbl["spec_detail"],
        "label_options":         lbl["options"],
        "label_phone":           lbl["phone"],
        "label_email":           lbl["email"],
        "label_web":             lbl["web"],
        "label_product_catalog": lbl["product_catalog"],
        "label_contact_us":      lbl["contact_us"],
        "label_address":         lbl["address"],
        "label_option_benefits": lbl["option_benefits"],
    })

    # WeasyPrint — also blocking, run in thread
    def _render_pdf():
        from weasyprint import HTML
        return HTML(string=html_str, base_url=base_url).write_pdf()

    try:
        pdf = await asyncio.to_thread(_render_pdf)
    except Exception as e:
        return Response(f"PDF oluşturulamadı: {e}", status_code=500, media_type="text/plain")

    normalized = unicodedata.normalize("NFKD", model_name)
    safe_name = "".join(c for c in normalized if ord(c) < 128)
    safe_name = safe_name.replace(" ", "_").replace("/", "-").strip("_")[:40] or "model"
    filename = f"katalog_{safe_name}_{lang}.pdf"
    return Response(
        pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/delete")
async def delete_model(request: Request, id: int = Form(...)):
    user = auth.require_user(request)
    if user["role"] != "admin":
        return RedirectResponse("/models?msg=Yetkisiz+işlem&msg_type=error", 303)
    fdb.del_model(id)
    return RedirectResponse("/models", 303)


@router.post("/request-delete")
async def request_delete_model(request: Request, id: int = Form(...)):
    user = auth.require_user(request)
    m = fdb.get_model(id)
    if not m:
        return RedirectResponse("/models?msg=Model+bulunamadı&msg_type=error", 303)
    group_ids = udb.get_manufacturer_group_ids(user)
    allowed_ids = set(int(x) for x in (user.get("allowed_models") or "").split(",") if x.strip().isdigit())
    if id not in allowed_ids and m.get("manufacturer_id") not in group_ids:
        return RedirectResponse("/models?msg=Yetkisiz+işlem&msg_type=error", 303)
    fdb.add_deletion_request("model", id, m["name"], user["id"])
    return RedirectResponse("/models?msg=Silme+talebi+yöneticiye+gönderildi&msg_type=success", 303)


@router.get("/excel-template")
async def excel_template_download(request: Request):
    auth.require_user(request)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response("openpyxl kurulu değil. pip install openpyxl", status_code=500, media_type="text/plain")

    wb = openpyxl.Workbook()
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _hdr(ws, col, value, fill_color, width, wrap=True):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font      = hdr_font
        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = width
        return cell

    # ── Sheet 1: Model temel bilgileri ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Model Bilgileri"
    ws1.row_dimensions[1].height = 36

    info = [
        ("Makine Adı (TR) *", "2563EB", 28),
        ("Makine Adı (EN)",   "2563EB", 28),
        ("Makine Adı (ZH)",   "2563EB", 28),
        ("Açıklama (TR)",     "2563EB", 35),
        ("Açıklama (EN)",     "2563EB", 35),
        ("Açıklama (ZH)",     "2563EB", 35),
        ("Satış Fiyatı",      "2563EB", 14),
        ("Para Birimi\n(USD/EUR/TRY)", "2563EB", 16),
        ("Kategori Adı",      "2563EB", 22),
    ]
    for col, (title, color, width) in enumerate(info, 1):
        _hdr(ws1, col, title, color, width)

    # Not satırı
    note_cell = ws1.cell(row=2, column=1, value="⚠ Resimler Excel'e eklenemez — formdaki resim alanından yüklenir.")
    note_cell.font = Font(italic=True, color="B45309")
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    # Örnek veri satırı
    example = ["Örnek Makine A", "Example Machine A", "", "Türkçe açıklama...", "English description...", "", 15000.00, "USD", "Kategori Adı"]
    for col, v in enumerate(example, 1):
        ws1.cell(row=3, column=col, value=v)

    # ── Sheet 2: Teknik özellikler ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Teknik Özellikler")
    ws2.row_dimensions[1].height = 28

    spec_info = [
        ("Başlık",          "059669", 30),
        ("Açıklama (TR)",   "059669", 45),
        ("Açıklama (EN)",   "059669", 45),
        ("Açıklama (ZH)",   "059669", 45),
    ]
    for col, (title, color, width) in enumerate(spec_info, 1):
        _hdr(ws2, col, title, color, width)

    examples2 = [
        ("Motor Gücü",      "3.5 kW / 4.8 HP",    "Motor Power: 3.5 kW / 4.8 HP", ""),
        ("Kesim Genişliği", "1600 mm",             "Cutting Width: 1600 mm",        ""),
        ("Tabla Boyutu",    "1600 x 1250 mm",      "Table Size: 1600 x 1250 mm",    ""),
        ("Max. Kesim Hızı", "60 m/min",            "Max. Cutting Speed: 60 m/min",  ""),
        ("Ağırlık",         "850 kg",              "Weight: 850 kg",                ""),
    ]
    for row_idx, row_data in enumerate(examples2, 2):
        for col, v in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col, value=v)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="model_sablonu.xlsx"'},
    )


@router.post("/excel-import")
async def excel_import(request: Request, excel_file: UploadFile = File(...)):
    auth.require_user(request)
    try:
        import openpyxl
    except ImportError:
        return JSONResponse({"error": "openpyxl kurulu değil"}, status_code=500)

    content = await excel_file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        return JSONResponse({"error": f"Dosya okunamadı: {e}"}, status_code=400)

    result = {}

    # ── Sheet 1: Model bilgileri ──────────────────────────────────────────────
    ws1 = wb[wb.sheetnames[0]]
    # Row 2 = note/merge, row 3 = actual data
    data_row_idx = 3 if ws1.max_row >= 3 else 2
    keys1 = ["name", "name_en", "name_zh", "description", "description_en", "description_zh",
             "base_price", "currency", "category_name"]
    row_vals = [cell.value for cell in ws1[data_row_idx]]
    for k, v in zip(keys1, row_vals):
        if v is None:
            continue
        if k == "base_price":
            try:
                result[k] = float(v)
            except Exception:
                result[k] = 0.0
        else:
            s = str(v).strip()
            if s:
                result[k] = s

    # ── Sheet 2: Teknik özellikler (Başlık | TR | EN | ZH) ────────────────────
    def _sv(v):
        return str(v).strip() if v is not None else ""

    specs, specs_en, specs_zh = [], [], []
    if len(wb.sheetnames) > 1:
        ws2 = wb[wb.sheetnames[1]]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            title = _sv(row[0])
            if not title:
                continue
            desc    = _sv(row[1]) if len(row) > 1 else ""
            desc_en = _sv(row[2]) if len(row) > 2 else ""
            desc_zh = _sv(row[3]) if len(row) > 3 else ""
            specs.append(   {"title": title, "desc": desc,    "img": ""})
            specs_en.append({"title": title, "desc": desc_en, "img": ""})
            specs_zh.append({"title": title, "desc": desc_zh, "img": ""})

    result["specs"]    = specs
    result["specs_en"] = specs_en
    result["specs_zh"] = specs_zh

    return JSONResponse(result)
