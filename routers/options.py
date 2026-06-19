import os
import uuid
from io import BytesIO
from typing import Optional, List

from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import RedirectResponse, Response

import db.factory as fdb
import db.users as udb
import auth
from config import CURRENCIES, QTY_TYPES, OPTION_SCOPES, IMAGES_DIR

router = APIRouter(prefix="/options")
from tmpl import templates


@router.get("")
async def options_list(request: Request, category_id: int = 0):
    user = auth.require_user(request)
    is_mfr = auth.is_mfr(user)
    mfr_filter = udb.get_manufacturer_group_ids(user) if is_mfr else None
    options = fdb.get_options(
        category_id=category_id if category_id else None,
        manufacturer_filter=mfr_filter,
    )
    cats = fdb.get_cats()

    # Apply allowed_categories restriction for manufacturers
    if is_mfr:
        # allowed_categories stored as names (e.g. "CNC İşleme,PVC Kenar Bantlama")
        allowed_cat_names = set(x.strip() for x in (user.get("allowed_categories") or "").split(",") if x.strip())
        if allowed_cat_names:
            allowed_cat_ids = {c["id"] for c in cats if c["name"] in allowed_cat_names}
            cats = [c for c in cats if c["id"] in allowed_cat_ids]
            options = [
                o for o in options
                if not (o.get("category_ids") or "").strip()  # uncategorized = visible to all
                or any(
                    int(x) in allowed_cat_ids
                    for x in (o.get("category_ids") or "").split(",")
                    if x.strip().isdigit()
                )
            ]

    cat_map = {c["id"]: c["name"] for c in cats}
    for o in options:
        cids = [int(x) for x in (o.get("category_ids") or "").split(",") if x.strip().isdigit()]
        o["category_names"] = ", ".join(cat_map[c] for c in cids if c in cat_map) or "Genel"
    manufacturers = udb.all_manufacturers() if user["role"] == "admin" else []
    return templates.TemplateResponse(request, "options.html", {
        "user": user,
        "options": options,
        "categories": cats,
        "selected_cat": category_id,
        "currencies": CURRENCIES,
        "qty_types": QTY_TYPES,
        "option_scopes": OPTION_SCOPES,
        "manufacturers": manufacturers,
        "active_page": "options",
        "msg": request.query_params.get("msg", ""),
        "msg_type": request.query_params.get("msg_type", "info"),
    })


@router.post("/save")
async def save_option(request: Request,
                      id: int = Form(0),
                      name: str = Form(...),
                      description: str = Form(""),
                      price: float = Form(0.0),
                      currency: str = Form("USD"),
                      scope: str = Form("GLOBAL"),
                      category_ids: List[int] = Form([]),
                      qty_type: str = Form("MANUAL"),
                      conflict_group: str = Form(""),
                      image_priority: int = Form(0),
                      current_image_path: str = Form(""),
                      current_variation_image_path: str = Form(""),
                      image: Optional[UploadFile] = File(None),
                      variation_image: Optional[UploadFile] = File(None),
                      name_en: str = Form(""),
                      description_en: str = Form(""),
                      name_zh: str = Form(""),
                      description_zh: str = Form(""),
                      video_url: str = Form(""),
                      manufacturer_id: int = Form(0),
                      requires_option_ids: List[int] = Form([]),
                      purchase_price: float = Form(0.0),
                      purchase_currency: str = Form("USD")):
    user = auth.require_user(request)

    def _save_file(upload: UploadFile, prefix: str) -> str:
        raise NotImplementedError

    async def _save(upload, prefix):
        if not (upload and upload.filename):
            return None
        os.makedirs(IMAGES_DIR, exist_ok=True)
        ext = os.path.splitext(upload.filename)[1].lower() or ".jpg"
        fname = f"{prefix}_{uuid.uuid4().hex[:10]}{ext}"
        content = await upload.read()
        with open(os.path.join(IMAGES_DIR, fname), "wb") as f:
            f.write(content)
        return f"img/uploads/{fname}"

    try:
        image_path = (await _save(image, "opt")) or current_image_path
        variation_image_path = (await _save(variation_image, "var")) or current_variation_image_path
    except Exception as e:
        return RedirectResponse(f"/options?msg=Resim+kaydedilemedi:+{e}&msg_type=error", 303)

    # Manufacturers automatically own their options; can only change purchase_price if currently 0
    if auth.is_mfr(user):
        price = 0.0
        manufacturer_id = udb.effective_mfr_id(user)
        if id:  # editing existing — only allow price change if price not yet set
            existing = fdb.get_option(id)
            existing_price = float(existing.get("purchase_price") or 0) if existing else 0.0
            if existing_price > 0:
                purchase_price = existing_price
                purchase_currency = existing.get("purchase_currency", "USD") if existing else "USD"

    kw = dict(
        name=name, description=description, price=price, currency=currency,
        scope=scope, category_ids=",".join(str(c) for c in category_ids), qty_type=qty_type,
        conflict_group=conflict_group, image_priority=image_priority,
        image_path=image_path, variation_image_path=variation_image_path,
        video_url=video_url.strip(),
        name_en=name_en, description_en=description_en,
        name_zh=name_zh, description_zh=description_zh,
        manufacturer_id=manufacturer_id or None,
        requires_option_ids=",".join(str(i) for i in requires_option_ids),
        purchase_price=purchase_price,
        purchase_currency=purchase_currency,
    )
    try:
        if id:
            # manufacturers can edit options they created or options assigned to them
            if auth.is_mfr(user):
                import db.users as udb
                existing = fdb.get_option(id)
                mfr_id = udb.effective_mfr_id(user)
                if not existing or (existing.get("created_by") != user["id"] and existing.get("manufacturer_id") != mfr_id):
                    return RedirectResponse("/options?msg=Yetkisiz+işlem&msg_type=error", 303)
            fdb.upd_option(id, **kw)
        else:
            kw["created_by"] = user["id"]
            fdb.add_option(**kw)
    except Exception as e:
        return RedirectResponse(f"/options?msg=Kayıt+hatası:+{e}&msg_type=error", 303)
    return RedirectResponse("/options?msg=Kaydedildi&msg_type=success", 303)


@router.post("/delete")
async def delete_option(request: Request, id: int = Form(...)):
    user = auth.require_user(request)
    if user["role"] != "admin":
        return RedirectResponse("/options?msg=Yetkisiz+işlem&msg_type=error", 303)
    fdb.del_option(id)
    return RedirectResponse("/options", 303)


@router.post("/request-delete")
async def request_delete_option(request: Request, id: int = Form(...)):
    user = auth.require_user(request)
    opt = fdb.get_option(id)
    if not opt:
        return RedirectResponse("/options?msg=Opsiyon+bulunamadı&msg_type=error", 303)
    group_ids = udb.get_manufacturer_group_ids(user)
    if opt.get("manufacturer_id") not in group_ids and opt.get("created_by") not in group_ids:
        return RedirectResponse("/options?msg=Yetkisiz+işlem&msg_type=error", 303)
    fdb.add_deletion_request("option", id, opt["name"], user["id"])
    return RedirectResponse("/options?msg=Silme+talebi+yöneticiye+gönderildi&msg_type=success", 303)


@router.get("/excel-template")
async def options_excel_template(request: Request):
    auth.require_user(request)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response("openpyxl kurulu değil. pip install openpyxl", status_code=500, media_type="text/plain")

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")

    def _hdr(ws, col, value, color, width):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font      = hdr_font
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    ws = wb.active
    ws.title = "Opsiyonlar"
    ws.row_dimensions[1].height = 32

    cols = [
        ("Opsiyon Adı (TR) *", "7C3AED", 30),
        ("Opsiyon Adı (EN)",   "7C3AED", 30),
        ("Opsiyon Adı (ZH)",   "7C3AED", 30),
        ("Açıklama (TR)",      "7C3AED", 40),
        ("Açıklama (EN)",      "7C3AED", 40),
        ("Açıklama (ZH)",      "7C3AED", 40),
    ]
    for col, (title, color, width) in enumerate(cols, 1):
        _hdr(ws, col, title, color, width)

    # Not satırı
    note = ws.cell(row=2, column=1, value="⚠ Resimler, fiyat ve diğer ayarlar opsiyon kaydedildikten sonra düzenleme ekranından girilir.")
    note.font = Font(italic=True, color="B45309")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Örnek satırlar
    examples = [
        ("Otomatik Besleme",  "Automatic Feeder",  "", "Otomatik malzeme besleme sistemi", "Automatic material feeding system", ""),
        ("Lazer İşaretleme",  "Laser Marking",     "", "Lazer ile ürün işaretleme modülü", "Laser product marking module",      ""),
        ("Ekstra Bıçak Seti", "Extra Blade Set",   "", "Yedek bıçak takımı (5 adet)",      "Spare blade set (5 pcs)",           ""),
    ]
    for row_idx, row_data in enumerate(examples, 3):
        for col, v in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=v)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="opsiyon_sablonu.xlsx"'},
    )


@router.post("/excel-import")
async def options_excel_import(request: Request, excel_file: UploadFile = File(...)):
    auth.require_admin(request)
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/options?msg=openpyxl+kurulu+değil&msg_type=error", 303)

    content = await excel_file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        return RedirectResponse(f"/options?msg=Dosya+okunamadı:+{e}&msg_type=error", 303)

    def _sv(v):
        return str(v).strip() if v is not None else ""

    ws = wb[wb.sheetnames[0]]
    added = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or row[0] is None:
            continue
        name = _sv(row[0])
        if not name:
            continue

        name_en     = _sv(row[1])
        name_zh     = _sv(row[2])
        description = _sv(row[3])
        desc_en     = _sv(row[4])
        desc_zh     = _sv(row[5])

        try:
            fdb.add_option(
                name=name, name_en=name_en, name_zh=name_zh,
                description=description, description_en=desc_en, description_zh=desc_zh,
                price=0.0, currency="USD", qty_type="MANUAL",
                category_ids="", conflict_group="", video_url="",
                scope="GLOBAL", image_path="", variation_image_path="", image_priority=0,
            )
            added += 1
        except Exception as e:
            errors.append(f"Satır {row_idx}: {e}")

    if errors:
        msg = f"{added}+opsiyon+eklendi,+{len(errors)}+hata"
    else:
        msg = f"{added}+opsiyon+başarıyla+eklendi"
    msg_type = "success" if not errors else "warning"
    return RedirectResponse(f"/options?msg={msg}&msg_type={msg_type}", 303)
